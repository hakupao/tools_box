import csv
import os
import re
from datetime import date, datetime, time
from typing import Callable, Dict, Iterable, List, Optional, Sequence, Set

from openpyxl import load_workbook


class XlsxSheetSplitter:
    """Excel 工作表拆分处理器。"""

    INVALID_CHARS_PATTERN = re.compile(r'[<>:"/\\|?*]')
    CONTROL_CHARS_PATTERN = re.compile(r'[\x00-\x1f]')
    RESERVED_NAMES = {
        "CON",
        "PRN",
        "AUX",
        "NUL",
        "COM1",
        "COM2",
        "COM3",
        "COM4",
        "COM5",
        "COM6",
        "COM7",
        "COM8",
        "COM9",
        "LPT1",
        "LPT2",
        "LPT3",
        "LPT4",
        "LPT5",
        "LPT6",
        "LPT7",
        "LPT8",
        "LPT9",
    }

    def __init__(self, output_encoding: str = "utf-8-sig") -> None:
        self.output_encoding = output_encoding

    def split_file(
        self,
        input_file: str,
        output_path: Optional[str] = None,
        progress_callback: Optional[Callable[[int, int, str], None]] = None,
    ) -> Dict[str, object]:
        """将单个 Excel 文件拆分为多个 CSV（按工作表）。

        Args:
            input_file: 输入 Excel 文件路径（.xlsx）。
            output_path: 输出目录，None 时输出到原文件所在目录。
            progress_callback: 进度回调 (current, total, sheet_name)。

        Returns:
            dict: 包含 output_dir、output_files、errors、total_sheets、success。
        """
        if not os.path.isfile(input_file):
            raise FileNotFoundError(f"文件不存在: {input_file}")

        output_dir = output_path or os.path.dirname(input_file)
        os.makedirs(output_dir, exist_ok=True)

        workbook = load_workbook(input_file, read_only=True, data_only=True)
        sheet_names = list(workbook.sheetnames)
        total = len(sheet_names)

        used_names: Set[str] = set()
        output_files: List[str] = []
        sheet_outputs: List[Dict[str, str]] = []
        errors: List[str] = []

        try:
            for index, sheet_name in enumerate(sheet_names, 1):
                if progress_callback:
                    progress_callback(index, total, sheet_name)

                safe_name = self._sanitize_sheet_name(sheet_name)
                safe_name = self._make_unique_name(safe_name, used_names)
                output_file = os.path.join(output_dir, f"{safe_name}.csv")

                try:
                    worksheet = workbook[sheet_name]
                    rows = self._read_sheet_as_strings(worksheet.iter_rows())
                except Exception as exc:  # pylint: disable=broad-except
                    errors.append(f"{sheet_name}: {exc}")
                    rows = []

                try:
                    self._write_csv(output_file, rows)
                    output_files.append(output_file)
                    sheet_outputs.append({"sheet": sheet_name, "output_file": output_file})
                except Exception as exc:  # pylint: disable=broad-except
                    errors.append(f"{sheet_name}: {exc}")
        finally:
            workbook.close()

        return {
            "input_file": input_file,
            "output_dir": output_dir,
            "total_sheets": total,
            "output_files": output_files,
            "sheet_outputs": sheet_outputs,
            "errors": errors,
            "success": len(errors) == 0,
        }

    @classmethod
    def _sanitize_sheet_name(cls, name: str) -> str:
        sanitized = cls.INVALID_CHARS_PATTERN.sub("_", name)
        sanitized = cls.CONTROL_CHARS_PATTERN.sub("_", sanitized)
        sanitized = sanitized.rstrip(" .")

        if sanitized in {"", ".", ".."}:
            sanitized = "Sheet"

        if sanitized.upper() in cls.RESERVED_NAMES:
            sanitized = f"{sanitized}_"

        return sanitized

    @staticmethod
    def _make_unique_name(name: str, used_names: Set[str]) -> str:
        if name not in used_names:
            used_names.add(name)
            return name

        counter = 1
        while True:
            candidate = f"{name}_{counter}"
            if candidate not in used_names:
                used_names.add(candidate)
                return candidate
            counter += 1

    @staticmethod
    def _read_sheet_as_strings(rows: Iterable[Iterable[object]]) -> List[List[str]]:
        output_rows: List[List[str]] = []
        max_nonempty_col = -1

        for row in rows:
            string_row = [XlsxSheetSplitter._cell_to_string(cell) for cell in row]
            for idx in range(len(string_row) - 1, -1, -1):
                if string_row[idx] != "":
                    if idx > max_nonempty_col:
                        max_nonempty_col = idx
                    break
            output_rows.append(string_row)

        if max_nonempty_col >= 0:
            output_rows = [row[: max_nonempty_col + 1] for row in output_rows]

        while output_rows and all(cell == "" for cell in output_rows[-1]):
            output_rows.pop()

        return output_rows

    def _write_csv(self, output_file: str, rows: List[List[str]]) -> None:
        with open(output_file, "w", newline="", encoding=self.output_encoding) as handle:
            writer = csv.writer(handle)
            writer.writerows(rows)

    @staticmethod
    def _cell_to_string(cell: object) -> str:
        if cell is None:
            return ""

        value = getattr(cell, "value", cell)
        if value is None:
            return ""

        if getattr(cell, "is_date", False) and isinstance(value, (datetime, date)):
            number_format = getattr(cell, "number_format", "")
            return XlsxSheetSplitter._format_excel_date(value, number_format)

        return str(value)

    @staticmethod
    def _format_excel_date(value: object, number_format: str) -> str:
        fmt = (number_format or "").lower()
        fmt = fmt.split(";")[0]
        fmt = re.sub(r"\[[^\]]*\]", "", fmt)
        fmt = re.sub(r'"[^"]*"', "", fmt)

        if "h" in fmt:
            date_part, time_part = fmt.split("h", 1)
            time_part = "h" + time_part
        else:
            date_part, time_part = fmt, ""

        date_tokens = XlsxSheetSplitter._extract_date_tokens(date_part)
        sep = XlsxSheetSplitter._detect_date_separator(date_part)
        date_str = XlsxSheetSplitter._render_date_tokens(value, date_tokens, sep)

        include_time = XlsxSheetSplitter._should_include_time(value, time_part)
        if not include_time:
            return date_str

        time_str = XlsxSheetSplitter._render_time(value, time_part)
        return f"{date_str} {time_str}"

    @staticmethod
    def _extract_date_tokens(date_part: str) -> Sequence[str]:
        tokens: List[str] = []
        i = 0
        while i < len(date_part):
            ch = date_part[i]
            if ch in {"y", "m", "d"}:
                j = i + 1
                while j < len(date_part) and date_part[j] == ch:
                    j += 1
                tokens.append(date_part[i:j])
                i = j
            else:
                i += 1
        return tokens

    @staticmethod
    def _detect_date_separator(date_part: str) -> str:
        if "/" in date_part:
            return "/"
        if "-" in date_part:
            return "-"
        if "." in date_part:
            return "."
        return "-"

    @staticmethod
    def _render_date_tokens(value: object, tokens: Sequence[str], sep: str) -> str:
        if not isinstance(value, (datetime, date)):
            return str(value)

        year = value.year
        month = value.month
        day = value.day

        if not tokens:
            return f"{year:04d}{sep}{month:02d}{sep}{day:02d}"

        parts: List[str] = []
        for token in tokens:
            if token.startswith("y"):
                if len(token) <= 2:
                    parts.append(f"{year % 100:02d}")
                else:
                    parts.append(f"{year:04d}")
            elif token.startswith("m"):
                if len(token) >= 2:
                    parts.append(f"{month:02d}")
                else:
                    parts.append(str(month))
            elif token.startswith("d"):
                if len(token) >= 2:
                    parts.append(f"{day:02d}")
                else:
                    parts.append(str(day))

        return sep.join(parts)

    @staticmethod
    def _should_include_time(value: object, time_part: str) -> bool:
        if not isinstance(value, datetime):
            return False
        if value.time() != time(0, 0, 0):
            return True
        return bool(time_part and ("h" in time_part or "s" in time_part))

    @staticmethod
    def _render_time(value: datetime, time_part: str) -> str:
        include_seconds = "s" in time_part
        return value.strftime("%H:%M:%S" if include_seconds else "%H:%M")
